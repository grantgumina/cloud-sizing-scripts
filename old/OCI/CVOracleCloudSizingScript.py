import sys
import subprocess
import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime
import logging
import tempfile
import json
import shutil

oci_path = shutil.which("oci")
kubectl_path = shutil.which("kubectl")

total_instances = 0
total_instance_sizeGB = 0
total_instance_sizeTB = 0
total_namespaces = 0
total_buckets = 0
total_storageGB = 0
total_storageTB = 0
total_db_systems = 0
total_db_system_sizeGB = 0
total_db_system_sizeTB = 0
total_oke_clusters = 0
total_oke_node_count = 0
total_oke_pvc_count = 0
total_oke_pvc_size_gb = 0
total_oke_pvc_size_tb = 0

class InstanceInfo:
    def __init__(self):
        self.compartment_id = None
        self.instance_id = None
        self.instance_name = None
        self.region = None
        self.availability_domain = None
        self.shape = None
        self.state = None
        self.number_of_volumes = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}
        self.boot_volume_name = None
        self.block_volume_names = []

class InstanceSummary:
    def __init__(self):
        self.region = None
        self.instance_count = 0
        self.total_sizeGB = 0
        self.total_sizeTB = 0

class ObjectStorageInfo:
    def __init__(self):
        self.compartment_id = None
        self.namespace = None
        self.bucket_name = None
        self.region = None
        self.storage_tier = None
        self.object_count = 0
        self.sizeGB = 0
        self.sizeTB = 0
        self.defined_tags = {}
        self.freeform_tags = {}

class ObjectStorageSummary:
    def __init__(self):
        self.region = None
        self.namespace = None
        self.bucket_count = 0
        self.total_storage_GB = 0
        self.total_storage_TB = 0

class DBSystemInfo:
    def __init__(self):
        self.compartment_id = None
        self.db_system_id = None
        self.display_name = None
        self.region = None
        self.availability_domain = None
        self.shape = None
        self.lifecycle_state = None
        self.node_count = 0
        self.db_version = None
        self.database_edition = None
        self.data_storage_size_gb = 0
        self.data_storage_size_tb = 0
        self.defined_tags = {}
        self.freeform_tags = {}

class DBSystemSummary:
    def __init__(self):
        self.region = None
        self.db_system_count = 0
        self.total_storage_gb = 0
        self.total_storage_tb = 0

class OKEClusterInfo:
    def __init__(self):
        self.region = None
        self.cluster_id = None
        self.cluster_name = None
        self.kubernetes_version = None
        self.node_count = 0
        self.node_names = []
        self.pvc_count = 0
        self.pvc_names = []
        self.total_pvc_size_gb = 0
        self.total_pvc_size_tb = 0

class OKEClusterSummary:
    def __init__(self):
        self.region = None
        self.cluster_count = 0
        self.total_node_count = 0
        self.total_pvc_count = 0
        self.total_pvc_size_gb = 0
        self.total_pvc_size_tb = 0


def install_and_import(package):
    try:
        __import__(package)
        print(f"Package '{package}' is already installed.")
    except ImportError:
        print(f"Package '{package}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def get_sheet_info(workload):
    if workload == "instances":
        info_sheet = "Instance Info"
        summary_sheet = "Instance Summary"
        info_headers = [
            "Compartment ID", "Instance ID", "Instance Name", "Region",
            "Availability Domain", "Shape", "State", "Number of Volumes",
            "Size (GB)", "Size (TB)", "Boot Volume Name", "Block Volume Names", "Defined Tags", "Freeform Tags"
        ]
        summary_headers = ["Region", "Instance Count", "Total Size (GB)", "Total Size (TB)"]
    elif workload == "object_storage":
        info_sheet = "Object Storage Info"
        summary_sheet = "Object Storage Summary"
        info_headers = [
            "Namespace", "Compartment ID", "Bucket Name", "Region",
            "Storage Tier", "Object Count", "Size (GB)", "Size (TB)",
            "Defined Tags", "Freeform Tags"
        ]
        summary_headers = ["Region", "Bucket Count", "Total Size (GB)", "Total Size (TB)"]
    elif workload == "db_systems":
        info_sheet = "DB System Info"
        summary_sheet = "DB System Summary"
        info_headers = [
            "Compartment ID", "DB System ID", "Display Name", "Region",
            "Availability Domain", "Shape", "Lifecycle State", "Node Count",
            "DB Version", "Database Edition", "Data Storage Size (GB)", "Data Storage Size (TB)",
            "Defined Tags", "Freeform Tags"
        ]
        summary_headers = ["Region", "DB System Count", "Total Storage (GB)", "Total Storage (TB)"]
    elif workload == "oke_clusters":
        info_sheet = "OKE Cluster Info"
        summary_sheet = "OKE Cluster Summary"
        info_headers = [
            "Region", "Cluster ID", "Cluster Name", "Kubernetes Version",
            "Node Count", "PVC Count", "Total PVC Size (GB)", "Total PVC Size (TB)",
            "PVC Names",  "Node Names"
        ]
        summary_headers = [
            "Region", "Cluster Count", "Total Node Count",
            "Total PVC Count", "Total PVC Size (GB)", "Total PVC Size (TB)"
        ]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    return info_sheet, summary_sheet, info_headers, summary_headers

def format_workbook(filename):
    wb = load_workbook(filename)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for sheet in wb.worksheets:
        # Format header row
        for cell in sheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col].width = adjusted_width
    wb.save(filename)

def init_excel(filename, workload):
    if not os.path.exists(filename):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        info_sheet, summary_sheet, info_headers, summary_headers = get_sheet_info(workload)

        wb.create_sheet(summary_sheet)
        wb.create_sheet(info_sheet)

        wb[info_sheet].append(info_headers)
        wb[summary_sheet].append(summary_headers)
        wb.save(filename)

def dump_info(filename, workload, object_list: list):
    info_sheet, _, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[info_sheet]

    for obj in object_list:
        if workload == "instances":
            row = [
                obj.compartment_id,
                obj.instance_id,
                obj.instance_name,
                obj.region,
                obj.availability_domain,
                obj.shape,
                obj.state,
                obj.number_of_volumes,
                obj.sizeGB,
                obj.sizeTB,
                obj.boot_volume_name if obj.boot_volume_name else "",
                ", ".join(obj.block_volume_names) if obj.block_volume_names else "",
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "object_storage":
            row = [
                obj.namespace,
                obj.compartment_id,
                obj.bucket_name,
                obj.region,
                obj.storage_tier,
                obj.object_count,
                obj.sizeGB,
                obj.sizeTB,
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "db_systems":
            row = [
                obj.compartment_id,
                obj.db_system_id,
                obj.display_name,
                obj.region,
                obj.availability_domain,
                obj.shape,
                obj.lifecycle_state,
                obj.node_count,
                obj.db_version,
                obj.database_edition,
                obj.data_storage_size_gb,
                obj.data_storage_size_tb,
                str(obj.defined_tags),
                str(obj.freeform_tags),
            ]
        elif workload == "oke_clusters":
            row = [
                obj.region,
                obj.cluster_id,
                obj.cluster_name,
                obj.kubernetes_version,
                obj.node_count,
                obj.pvc_count,
                obj.total_pvc_size_gb,
                obj.total_pvc_size_tb,
                ", ".join(obj.pvc_names) if obj.pvc_names else "",
                ", ".join(obj.node_names) if obj.node_names else "",
            ]
        else:
            raise ValueError(f"Unsupported workload: {workload}")
        sheet.append(row)
    wb.save(filename)

def dump_summary(filename, workload, summary):
    _, summary_sheet, _, _ = get_sheet_info(workload)
    wb = load_workbook(filename)
    sheet = wb[summary_sheet]

    for obj in summary:
        if workload == "instances":
            row = [
                obj.region,
                obj.instance_count,
                obj.total_sizeGB,
                obj.total_sizeTB,
            ]
        elif workload == "object_storage":
            row = [
                obj.region,
                obj.bucket_count,
                obj.total_storage_GB,
                obj.total_storage_TB,
            ]
        elif workload == "db_systems":
            row = [
                obj.region,
                obj.db_system_count,
                obj.total_storage_gb,
                obj.total_storage_tb,
            ]
        elif workload == "oke_clusters":
            row = [
                obj.region,
                obj.cluster_count,
                obj.total_node_count,
                obj.total_pvc_count,
                obj.total_pvc_size_gb,
                obj.total_pvc_size_tb
            ]
        else:
            raise ValueError(f"Unsupported workload: {workload}")
        sheet.append(row)

    wb.save(filename)

def write_grand_total(filename, workload):
    wb = load_workbook(filename)

    bold_font = Font(bold=True)
    if workload == "instances":
        global total_instances, total_instance_sizeGB, total_instance_sizeTB
        sheet_name = "Instance Summary"
        row = ["Total Instances", total_instances, total_instance_sizeGB, total_instance_sizeTB]
    elif workload == "object_storage":
        sheet_name = "Object Storage Summary"
        global total_namespaces, total_buckets, total_storageGB, total_storageTB
        row = ["Total Buckets", total_buckets, total_storageGB, total_storageTB]
    elif workload == "db_systems":
        sheet_name = "DB System Summary"
        global total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB
        row = ["Total DB Systems", total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB]
    elif workload == "oke_clusters":
        sheet_name = "OKE Cluster Summary"
        global total_oke_clusters, total_oke_node_count, total_oke_pvc_count
        global total_oke_pvc_size_gb, total_oke_pvc_size_tb
        row = [
            "Total OKE Clusters",
            total_oke_clusters,
            total_oke_node_count,
            total_oke_pvc_count,
            total_oke_pvc_size_gb,
            total_oke_pvc_size_tb
        ]
    else:
        raise ValueError(f"Unsupported workload: {workload}")

    sheet = wb[sheet_name]
    sheet.append(row)
    for cell in sheet[2]:
        cell.font = bold_font
    wb.save(filename)

def get_object_storage_info(config, filename, regions=[], compartments=[]):
    global total_buckets, total_storageGB, total_storageTB
    object_storage_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        object_storage_client = oci.object_storage.ObjectStorageClient(config)
        region_summary = ObjectStorageSummary()
        region_summary.region = region
        region_summary.bucket_count = 0
        region_summary.total_storage_GB = 0
        region_summary.total_storage_TB = 0
        try:
            namespace = object_storage_client.get_namespace().data
        except Exception as e:
            logging.error(f"Error fetching namespace for region {region}: {e}")
            continue
        for compartment in compartments:
            compartment_bucket_list = []
            try:
                buckets = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_buckets,
                    namespace_name=namespace,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching buckets for compartment {compartment}: {e}")
                continue
            logging.info(f"Found {len(buckets)} bucket(s) in compartment {compartment}")
            if len(buckets) == 0:
                continue
            for bucket in buckets:
                bucket_info = ObjectStorageInfo()
                bucket_info.compartment_id = compartment
                bucket_info.namespace = namespace
                bucket_info.bucket_name = bucket.name
                bucket_info.region = region
                try:
                    stats = object_storage_client.get_bucket(
                        namespace_name=namespace,
                        bucket_name=bucket.name,
                        fields=['approximateSize', 'approximateCount']
                    ).data
                    bucket_info.storage_tier = stats.storage_tier
                    bucket_info.defined_tags = stats.defined_tags
                    bucket_info.freeform_tags = stats.freeform_tags
                    size_in_bytes, object_count = stats.approximate_size, stats.approximate_count
                    bucket_info.sizeGB = round(size_in_bytes / (1024 ** 3), 2) if size_in_bytes else 0
                    bucket_info.sizeTB = round(bucket_info.sizeGB / 1024, 2) if bucket_info.sizeGB else 0
                    bucket_info.object_count = object_count
                except Exception as e:
                    logging.error(f"Error fetching stats for bucket {bucket.name}: {e}")
                    continue
                region_summary.bucket_count += 1
                region_summary.total_storage_GB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                region_summary.total_storage_TB += bucket_info.sizeTB if bucket_info.sizeTB else 0
                compartment_bucket_list.append(bucket_info)
                total_buckets += 1
                total_storageGB += bucket_info.sizeGB if bucket_info.sizeGB else 0
                total_storageTB += bucket_info.sizeTB if bucket_info.sizeTB else 0
            dump_info(filename, "object_storage", compartment_bucket_list)
        object_storage_summary_list.append(region_summary)
        
    write_grand_total(filename, "object_storage")
    dump_summary(filename, "object_storage", object_storage_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments for object storage.")
    logging.info(f"Grand Total - Buckets: {total_buckets}, Size (GB): {total_storageGB}, Size (TB): {total_storageTB}")

def get_boot_volume_info(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_boot_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        boot_volumes = response.data
        result = {"name": None, "sizeGB": 0}
        if not boot_volumes:
            return result
        try:
            response = block_storage_client.get_boot_volume(boot_volumes[0].boot_volume_id)
            boot_volume_info = response.data
            result = {"name": boot_volume_info.display_name, "sizeGB": boot_volume_info.size_in_gbs}
            return result
        except Exception as e:
            print(f"Error retrieving boot volume info for instance {instance_id}: {e}")
            return result
    except Exception as e:
        print(f"Error retrieving boot volume attachments for instance {instance_id}: {e}")
        return {"name": None, "sizeGB": 0}

def get_block_volume_info(config, instance_id, availability_domain, compartment_id):
    compute_client = oci.core.ComputeClient(config)
    block_storage_client = oci.core.BlockstorageClient(config)
    try:
        response = oci.pagination.list_call_get_all_results(compute_client.list_volume_attachments,
                                                            instance_id=instance_id,
                                                            availability_domain=availability_domain,
                                                            compartment_id=compartment_id,
                                                            retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY)
        volume_attachments = response.data
        result = []
        if not volume_attachments:
            return result
        for attachment in volume_attachments:
            try:
                response = block_storage_client.get_volume(attachment.volume_id)
                volume_info = response.data
                result.append({"name": volume_info.display_name, "sizeGB": volume_info.size_in_gbs})
            except Exception as e:
                print(f"Error retrieving volume info for volume {attachment.volume_id}: {e}")
        return result
    except Exception as e:
        print(f"Error retrieving volume attachments for instance {instance_id}: {e}")
        return []

def get_instance_info(config, filename, regions=[], compartments=[]):
    global total_instances, total_instance_sizeGB, total_instance_sizeTB
    instance_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions: 
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        compute_client = oci.core.ComputeClient(config)
        region_summary = InstanceSummary()
        for compartment in compartments:
            logging.info(f"Processing compartment: {compartment}")
            compartment_instance_list = []
            region_summary.region = region
            instances = oci.pagination.list_call_get_all_results(compute_client.list_instances,
                                                                compartment_id=compartment,
                                                                retry_strategy=oci.retry.DEFAULT_RETRY_STRATEGY).data
            logging.info(f"Found {len(instances)} instance(s)")
            if len(instances) == 0:
                continue
            for instance in instances:
                if instance.lifecycle_state == "TERMINATED":
                    continue
                logging.info(f"Processing instance: {instance.id} - {instance.display_name}")
                instance_info = InstanceInfo()
                instance_info.compartment_id = compartment
                instance_info.instance_id = instance.id
                instance_info.instance_name = instance.display_name
                instance_info.region = region
                instance_info.availability_domain = instance.availability_domain
                instance_info.shape = instance.shape
                instance_info.state = instance.lifecycle_state
                instance_info.defined_tags = instance.defined_tags
                instance_info.freeform_tags = instance.freeform_tags
                try:
                    boot_volume_info = get_boot_volume_info(config, instance.id, instance.availability_domain, instance.compartment_id)
                    block_volumes_info = get_block_volume_info(config, instance.id, instance.availability_domain, instance.compartment_id)
                except Exception as e:
                    logging.error(f"Error fetching volume data for instance {instance.id}: {e}")
                    continue
                instance_info.number_of_volumes = (1 if boot_volume_info["sizeGB"] > 0 else 0) + len(block_volumes_info)
                instance_info.sizeGB = boot_volume_info["sizeGB"] + sum([bv["sizeGB"] for bv in block_volumes_info])
                instance_info.sizeTB = round(instance_info.sizeGB / 1024, 2)
                instance_info.boot_volume_name = boot_volume_info["name"] if boot_volume_info["name"] else None
                instance_info.block_volume_names = [bv["name"] for bv in block_volumes_info]
                region_summary.instance_count += 1
                region_summary.total_sizeGB += instance_info.sizeGB
                region_summary.total_sizeTB += instance_info.sizeTB
                compartment_instance_list.append(instance_info)
                total_instances += 1
                total_instance_sizeGB += instance_info.sizeGB
                total_instance_sizeTB += instance_info.sizeTB
            dump_info(filename, "instances", compartment_instance_list)
        instance_summary_list.append(region_summary)
    write_grand_total(filename, "instances")
    dump_summary(filename, "instances", instance_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments.")
    logging.info(f"Grand Total - Instances: {total_instances}, Size (GB): {total_instance_sizeGB}, Size (TB): {total_instance_sizeTB}")

def get_database_info(config, filename, regions=[], compartments=[]):
    global total_db_systems, total_db_system_sizeGB, total_db_system_sizeTB
    db_summary_list = []
    identity_client = oci.identity.IdentityClient(config)
    if not regions:
        regions = [region.region_name for region in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [compartment.id for compartment in identity_client.list_compartments(compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]
    for region in regions:
        logging.info(f"Processing region: {region}")
        config["region"] = region
        db_client = oci.database.DatabaseClient(config)
        region_summary = DBSystemSummary()
        region_summary.region = region
        for compartment in compartments:
            logging.info(f"Processing compartment: {compartment}")
            compartment_db_list = []
            try:
                db_systems = oci.pagination.list_call_get_all_results(
                    db_client.list_db_systems,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching DB systems for compartment {compartment}: {e}")
                continue
            logging.info(f"Found {len(db_systems)} DB system(s)")
            if len(db_systems) == 0:
                continue
            for db in db_systems:
                if db.lifecycle_state == "TERMINATED":
                    continue
                db_info = DBSystemInfo()
                db_info.compartment_id = compartment
                db_info.db_system_id = db.id
                db_info.display_name = db.display_name
                db_info.region = region
                db_info.availability_domain = db.availability_domain
                db_info.shape = db.shape
                db_info.lifecycle_state = db.lifecycle_state
                db_info.node_count = db.node_count if hasattr(db, "node_count") else 0
                db_info.db_version = db.version if hasattr(db, "version") else ""
                db_info.database_edition = db.database_edition if hasattr(db, "database_edition") else "" 
                db_info.data_storage_size_gb = db.data_storage_size_in_gbs if hasattr(db, "data_storage_size_in_gbs") else 0
                db_info.data_storage_size_tb = round(db_info.data_storage_size_gb / 1024, 2)
                db_info.defined_tags = db.defined_tags
                db_info.freeform_tags = db.freeform_tags
                region_summary.db_system_count += 1
                region_summary.total_storage_gb += db_info.data_storage_size_gb
                region_summary.total_storage_tb += db_info.data_storage_size_tb
                compartment_db_list.append(db_info)
                total_db_systems += 1
                total_db_system_sizeGB += db_info.data_storage_size_gb
                total_db_system_sizeTB += db_info.data_storage_size_tb
            dump_info(filename, "db_systems", compartment_db_list)
        db_summary_list.append(region_summary)
    write_grand_total(filename, "db_systems")
    dump_summary(filename, "db_systems", db_summary_list)
    format_workbook(filename)
    logging.info("Completed processing all regions and compartments for DB systems.")
    logging.info(f"Grand Total - DB Systems: {total_db_systems}, Storage (GB): {total_db_system_sizeGB}, Storage (TB): {total_db_system_sizeTB}")

def get_oke_cluster_info(config, filename, regions=[], compartments=[]):
    global total_oke_clusters, total_oke_node_count, total_oke_pvc_count
    global total_oke_pvc_size_gb, total_oke_pvc_size_tb

    oke_summary_list = []
    identity_client = oci.identity.IdentityClient(config)

    if not regions:
        regions = [r.region_name for r in identity_client.list_region_subscriptions(config["tenancy"]).data]
    if not compartments:
        compartments = [c.id for c in identity_client.list_compartments(
            compartment_id=config["tenancy"], compartment_id_in_subtree=True).data]

    for region in regions:
        logging.info(f"Processing OKEs in region: {region}")
        config["region"] = region
        container_engine_client = oci.container_engine.ContainerEngineClient(config)

        region_summary = OKEClusterSummary()
        region_summary.region = region

        for compartment in compartments:
            try:
                clusters = oci.pagination.list_call_get_all_results(
                    container_engine_client.list_clusters,
                    compartment_id=compartment
                ).data
            except Exception as e:
                logging.error(f"Error fetching clusters in {compartment}: {e}")
                continue

            compartment_oke_list = []
            for cluster in clusters:
                if cluster.lifecycle_state == "DELETED":
                    continue

                cluster_info = OKEClusterInfo()
                cluster_info.region = region
                cluster_info.cluster_id = cluster.id
                cluster_info.cluster_name = cluster.name
                cluster_info.kubernetes_version = cluster.kubernetes_version

                kubeconfig_file = os.path.join(tempfile.gettempdir(), f"kubeconfig_{cluster.id}")
                try:
                    subprocess.run(
                        [
                            "oci", "ce", "cluster", "create-kubeconfig",
                            "--cluster-id", cluster.id,
                            "--file", kubeconfig_file,
                            "--region", region,
                            "--token-version", "2.0.0",
                            "--kube-endpoint", "PRIVATE_ENDPOINT",
                            "--profile", config.get("profile", oci.config.DEFAULT_PROFILE)
                        ],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    logging.info(f"Kubeconfig created at {kubeconfig_file} for cluster {cluster.name}")

                    result = subprocess.run(
                        ["kubectl", "--kubeconfig", kubeconfig_file, "get", "pvc", "-A", "-o", "json"],
                        check=True,
                        capture_output=True,
                        text=True
                    )

                    pvc_data = json.loads(result.stdout)
                    cluster_info.pvc_count = len([pvc for pvc in pvc_data["items"] if pvc.get("metadata", {}).get("name")])
                    cluster_info.pvc_names = [
                        f"{pvc['metadata'].get('namespace','default')}/{pvc['metadata']['name']}"
                        for pvc in pvc_data["items"]
                    ]
                    cluster_info.total_pvc_size_gb = sum([
                        int(pvc["spec"]["resources"]["requests"]["storage"].replace("Gi", ""))
                        for pvc in pvc_data["items"]
                        if "resources" in pvc["spec"] and "storage" in pvc["spec"]["resources"]["requests"]
                    ])
                    cluster_info.total_pvc_size_tb = round(cluster_info.total_pvc_size_gb / 1024, 2)

                    try:
                        node_result = subprocess.run(
                            ["kubectl", "--kubeconfig", kubeconfig_file, "get", "nodes", "-o", "json"],
                            check=True,
                            capture_output=True,
                            text=True
                        )
                        node_data = json.loads(node_result.stdout)
                        cluster_info.node_names = [n["metadata"]["name"] for n in node_data["items"]]
                        cluster_info.node_count = len(cluster_info.node_names)
                    except Exception as e:
                        logging.warning(f"Could not fetch nodes for {cluster.name}: {e}")
                        cluster_info.node_names = []
                        cluster_info.node_count = 0

                except Exception as e:
                    logging.warning(f"Error while using kubectl to fetch nodes and pvcs for {cluster.name}: {e}")
                finally:
                    try:
                        os.remove(kubeconfig_file)
                    except OSError:
                        pass


                region_summary.cluster_count += 1
                region_summary.total_node_count += cluster_info.node_count
                region_summary.total_pvc_count += cluster_info.pvc_count
                region_summary.total_pvc_size_gb += cluster_info.total_pvc_size_gb
                region_summary.total_pvc_size_tb += cluster_info.total_pvc_size_tb

                total_oke_clusters += 1
                total_oke_node_count += cluster_info.node_count
                total_oke_pvc_count += cluster_info.pvc_count
                total_oke_pvc_size_gb += cluster_info.total_pvc_size_gb
                total_oke_pvc_size_tb += cluster_info.total_pvc_size_tb

                compartment_oke_list.append(cluster_info)

            dump_info(filename, "oke_clusters", compartment_oke_list)

        oke_summary_list.append(region_summary)

    write_grand_total(filename, "oke_clusters")
    dump_summary(filename, "oke_clusters", oke_summary_list)
    format_workbook(filename)
    logging.info("Completed processing OKE clusters.")
    logging.info(
        f"Grand Total - OKE Clusters: {total_oke_clusters}, "
        f"Nodes: {total_oke_node_count}, PVCs: {total_oke_pvc_count}, "
        f"PVC Size (GB): {total_oke_pvc_size_gb}, PVC Size (TB): {total_oke_pvc_size_tb}"
    )


if __name__ == "__main__":

    if not shutil.which("kubectl"):
        logging.error("Error: 'kubectl' command not found. Please install kubectl to proceed.")
        sys.exit(1)
    if not shutil.which("oci"):
        logging.error("Error: 'oci' CLI not found. Please install OCI CLI to proceed.")
        sys.exit(1)

    packages = ["oci", "openpyxl", "pandas"]

    for pkg in packages:
        install_and_import(pkg)

    args = sys.argv[1:]
    for arg in args:
        if arg.startswith("--profile="):
            profile_name = arg.split("=")[1]
        elif arg.startswith("--region="):
            regions = arg.split("=")[1].split(",")
        elif arg.startswith("--compartment="):
            compartments = arg.split("=")[1].split(",")
        elif arg.startswith("--workload="):
            workload = arg.split("=")[1]
        elif arg == "--help":
            print("Usage: python oci_cs.py [--workload=<instances|object_storage|db_systems|oke_clusters>] [--profile=<profilename>] [--region=<region1>,<region2>] [--compartment=<comp1>,<comp2>] [--help]")
            sys.exit(0)
        else:
            print(f"Unknown argument: {arg}")
            sys.exit(1)

    if "workload" not in locals():
        workload = "all"
    if "profile_name" not in locals():
        profile_name = oci.config.DEFAULT_PROFILE
    if "regions" not in locals():
        regions = []
    if "compartments" not in locals():
        compartments = []
    config = oci.config.from_file(profile_name=profile_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    log_dir = "Logs"
    os.makedirs(log_dir, exist_ok=True)
    log_filename = os.path.join(log_dir, f"{profile_name}_{workload}_{timestamp}.log")
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
    logging.basicConfig(
        handlers=handlers,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    metrics_dir = "Metrics"
    os.makedirs(metrics_dir, exist_ok=True)
    filename = os.path.join(metrics_dir, f"{profile_name}_{workload}_{timestamp}.xlsx")
    if workload == "instances":
        init_excel(filename,workload)
        get_instance_info(config, filename, regions, compartments)
    elif workload == "object_storage":
        init_excel(filename,workload)
        get_object_storage_info(config, filename, regions, compartments)
    elif workload == "db_systems":
        init_excel(filename,workload)
        get_database_info(config, filename, regions, compartments)
    elif workload == "oke_clusters":
        init_excel(filename, workload)
        get_oke_cluster_info(config, filename, regions, compartments)
    elif workload == "all":
        logging.info(f"Getting information for all supported workloads.")
        for wl in ["instances", "object_storage", "db_systems", "oke_clusters"]:
            init_excel(filename, wl)
            if wl == "instances":
                get_instance_info(config, filename, regions, compartments)
            elif wl == "object_storage":
                get_object_storage_info(config, filename, regions, compartments)
            elif wl == "db_systems":
                get_database_info(config, filename, regions, compartments)
            elif wl == "oke_clusters":
                get_oke_cluster_info(config, filename, regions, compartments)
    else:
        logging.error(f"Unsupported workload specified: {workload}. Supported workloads are: instances, object_storage, db_systems, oke_clusters. If you want to gather information for all workloads, use --workload=all or don't specify the --workload argument at all.")
        sys.exit(1)