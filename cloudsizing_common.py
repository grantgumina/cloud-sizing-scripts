#!/usr/bin/env python3
"""
cloudsizing_common.py

Cloud-agnostic toolkit shared by the per-cloud sizing scripts (AWS, Azure, ...).

It owns everything that is NOT cloud-specific:
  * generic unit/byte/tag/k8s helpers,
  * the dataclass -> Excel workbook pipeline (build_summary / write_workbook /
    format_workbook), driven entirely by each workload dataclass's class
    attributes (WORKLOAD / *_SHEET / *_HEADERS / SUMMARY_SUM_FIELDS / GRAND_LABEL
    / to_row()),
  * the rich terminal UX (shared stderr Console, logging split, progress
    Reporter, summary table, JSON summary, banner),
  * shared CLI argument helpers,
  * a run_sizing(config) driver that owns the whole main() control flow and the
    top-level interrupt/error handling.

Each cloud script supplies a CloudConfig of hooks (auth, collect, parse_args,
...) plus its own workload dataclasses + collectors, then calls
run_sizing(CONFIG). Reference both clouds: AWS/CVAWSCloudSizingScript.py and
Azure/CVAzureCloudSizingScript.py.
"""

import sys
import os
import json
import re
import logging
from datetime import datetime
from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from rich.console import Console  # noqa: E402
from rich.logging import RichHandler  # noqa: E402
from rich.progress import (  # noqa: E402
    Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn,
)
from rich.table import Table  # noqa: E402
from rich.text import Text  # noqa: E402


# Single shared console for ALL human-facing output (status, progress, tables,
# logs) -> stderr. stdout is reserved for machine output (file paths / --json).
console = Console(stderr=True)

# Set by run_sizing once args are parsed; read by the entrypoint try/except to
# decide whether to show a full traceback.
_VERBOSE = False


def disable_color():
    """Reassign the shared console with color disabled (--no-color)."""
    global console
    console = Console(stderr=True, no_color=True)


# --------------------------------------------------------------------------- #
# Generic helpers (cloud-agnostic)
# --------------------------------------------------------------------------- #
def convert_bytes_to_sizes(num_bytes):
    """Return both binary (GiB/TiB) and decimal (GB/TB) sizes for a byte count."""
    num_bytes = num_bytes or 0
    return {
        "GiB": round(num_bytes / (1024 ** 3), 4),
        "TiB": round(num_bytes / (1024 ** 4), 4),
        "GB": round(num_bytes / 1e9, 4),
        "TB": round(num_bytes / 1e12, 4),
    }


def tags_to_dict(tag_list, key_field="Key", value_field="Value"):
    """Convert a [{Key, Value}, ...] style tag list to a plain dict.

    (Azure tags are already dicts; pass them through dict() instead.)
    """
    out = {}
    for t in tag_list or []:
        try:
            out[t[key_field]] = t[value_field]
        except (KeyError, TypeError):
            continue
    return out


def convert_k8s_size_to_bytes(qty):
    """Parse a Kubernetes resource quantity (e.g. '100Gi', '5G') to bytes."""
    if qty is None:
        return 0
    qty = str(qty).strip()
    if not qty:
        return 0
    binary = {"Ki": 1024, "Mi": 1024 ** 2, "Gi": 1024 ** 3,
              "Ti": 1024 ** 4, "Pi": 1024 ** 5, "Ei": 1024 ** 6}
    decimal = {"k": 1e3, "K": 1e3, "M": 1e6, "G": 1e9,
               "T": 1e12, "P": 1e15, "E": 1e18}
    m = re.match(r"^([0-9.]+)\s*([A-Za-z]+)?$", qty)
    if not m:
        return 0
    value = float(m.group(1))
    suffix = m.group(2) or ""
    if suffix in binary:
        return int(value * binary[suffix])
    if suffix in decimal:
        return int(value * decimal[suffix])
    return int(value)


# --------------------------------------------------------------------------- #
# Excel output (driven by each workload dataclass's class attributes)
# --------------------------------------------------------------------------- #
def build_summary(infos, cls):
    """Return (region_rows, grand_total_row) for a list of Info objects."""
    by_region = {}
    for o in infos:
        region = o.region or "unknown"
        bucket = by_region.setdefault(
            region, {"count": 0, "sums": {f: 0 for f, _ in cls.SUMMARY_SUM_FIELDS}})
        bucket["count"] += 1
        for f, _ in cls.SUMMARY_SUM_FIELDS:
            bucket["sums"][f] += getattr(o, f, 0) or 0

    region_rows = []
    for region in sorted(by_region):
        d = by_region[region]
        region_rows.append([region, d["count"]] +
                           [round(d["sums"][f], 4) for f, _ in cls.SUMMARY_SUM_FIELDS])

    total_count = sum(d["count"] for d in by_region.values())
    total_sums = [round(sum(d["sums"][f] for d in by_region.values()), 4)
                  for f, _ in cls.SUMMARY_SUM_FIELDS]
    grand = [cls.GRAND_LABEL, total_count] + total_sums
    return region_rows, grand


def write_workbook(filename, data_by_workload, workload_classes):
    """Write a workbook with an Info + Summary sheet pair per non-empty workload."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    for cls in workload_classes:
        infos = data_by_workload.get(cls.WORKLOAD)
        if not infos:
            continue

        summary_ws = wb.create_sheet(cls.SUMMARY_SHEET)
        summary_ws.append(cls.SUMMARY_HEADERS)
        region_rows, grand = build_summary(infos, cls)
        summary_ws.append(grand)           # grand total at row 2 (bolded below)
        for row in region_rows:
            summary_ws.append(row)
        for cell in summary_ws[2]:
            cell.font = Font(bold=True)

        info_ws = wb.create_sheet(cls.INFO_SHEET)
        info_ws.append(cls.INFO_HEADERS)
        for o in infos:
            info_ws.append(o.to_row())

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(filename)
    format_workbook(filename)


def format_workbook(filename):
    """Bold/shade header rows and auto-size columns."""
    wb = load_workbook(filename)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = bold_font
            cell.fill = header_fill
        for column_cells in sheet.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col].width = min(max_length + 2, 80)
    wb.save(filename)


# --------------------------------------------------------------------------- #
# Console UI: logging, progress reporter, summary table, JSON, banner
# --------------------------------------------------------------------------- #
def setup_logging(timestamp, log_prefix="sizing", quiet=False, verbose=False,
                  noisy_loggers=()):
    """Configure a full timestamped file log plus a clean rich console log.

    The file log keeps `asctime - level - message` for analysts; the console
    (stderr) shows human-friendly lines with no timestamp/level prefix unless
    --verbose. noisy_loggers are pinned to WARNING (e.g. SDK debug spam).
    Returns the log file path.
    """
    os.makedirs("Logs", exist_ok=True)
    log_filename = os.path.join("Logs", f"{log_prefix}_{timestamp}.log")

    file_handler = logging.FileHandler(log_filename)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    console_level = (logging.WARNING if quiet
                     else logging.DEBUG if verbose else logging.INFO)
    console_handler = RichHandler(
        console=console, level=console_level, markup=False,
        show_time=verbose, show_level=verbose, show_path=verbose,
        rich_tracebacks=True)

    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.handlers = [file_handler, console_handler]

    # Pin noisy SDK loggers to ERROR so their chatter stays out of our output
    # even under --verbose (e.g. azure-identity logs its full credential-chain
    # breakdown at WARNING; botocore/urllib3 are very verbose at DEBUG/INFO).
    for noisy in noisy_loggers:
        logging.getLogger(noisy).setLevel(logging.ERROR)
    return log_filename


class Reporter:
    """Drives a transient rich progress display for the collection run.

    No-ops when the console is not a TTY or when --quiet, so piped/CI output and
    --json stay clean. The collect hook drives start_account/update/advance; the
    `unit` noun labels the completed/total counter ("regions", "workloads", ...).
    """

    def __init__(self, enabled=True, unit="steps"):
        self.enabled = enabled and console.is_terminal
        self.unit = unit
        self.progress = None
        self.task = None

    def __enter__(self):
        if self.enabled:
            self.progress = Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TextColumn("{task.completed}/{task.total} " + self.unit),
                TimeElapsedColumn(),
                console=console, transient=True)
            self.progress.start()
        return self

    def __exit__(self, *exc):
        if self.progress:
            self.progress.stop()
        return False

    def start_account(self, label, total):
        if not self.progress:
            return
        if self.task is not None:
            self.progress.remove_task(self.task)
        self.task = self.progress.add_task(label, total=max(total, 1))

    def update(self, description):
        if self.progress and self.task is not None:
            self.progress.update(self.task, description=description)

    def advance(self):
        if self.progress and self.task is not None:
            self.progress.advance(self.task)


def _workload_totals(data, workload_classes):
    """Return {workload: {count, <sum fields>}} grand totals for a data dict."""
    out = {}
    for cls in workload_classes:
        infos = data.get(cls.WORKLOAD)
        if not infos:
            continue
        _, grand = build_summary(infos, cls)
        block = {"count": grand[1]}
        for (field_name, _header), value in zip(cls.SUMMARY_SUM_FIELDS, grand[2:]):
            block[field_name] = value
        out[cls.WORKLOAD] = block
    return out


def render_summary_table(combined, workload_classes, extra=None):
    """Print a per-workload grand-total table to the console (stderr).

    `extra` (optional) is {"title", "rows": [(label, value), ...]} appended as a
    labeled section below the workload rows (e.g. backup coverage)."""
    totals = _workload_totals(combined, workload_classes)
    if not totals:
        console.print("No resources found.", style="yellow")
        return
    table = Table(title="Sizing summary", title_style="bold",
                  header_style="bold", title_justify="left")
    table.add_column("Workload")
    table.add_column("Resource Count", justify="right")
    table.add_column("Total (TB)", justify="right")
    for cls in workload_classes:
        block = totals.get(cls.WORKLOAD)
        if not block:
            continue
        name = cls.INFO_SHEET.replace(" Info", "")
        tb = block.get("size_tb")
        tb_str = f"{tb:,.4f}" if tb is not None else "-"
        if cls.WORKLOAD == "cloud-rewind-quote":
            # Section header row for Cloud Rewind (no count/size values)
            header = Text(name, style="bold")
            table.add_row(header, "", "")
            # Sub-rows indented with tree characters
            discovered = block.get("total_count", 0)
            protectable = block.get("protectable_count", 0)
            table.add_row(
                Text("  ├─ Discovered Resources", style="dim"),
                f"{discovered:,}", "-",
            )
            table.add_row(
                Text("  └─ Protectable Resources", style="dim"),
                f"{protectable:,}", "-",
            )
        else:
            table.add_row(name, f"{block['count']:,}", tb_str)
    if extra and extra.get("rows"):
        table.add_row(Text(extra.get("title", "Extra"), style="bold"), "", "")
        rows = extra["rows"]
        for i, (lbl, val) in enumerate(rows):
            connector = "└─" if i == len(rows) - 1 else "├─"
            table.add_row(Text(f"  {connector} {lbl}", style="dim"), str(val), "")
    console.print(table)


def build_json_summary(scope_items, combined, comprehensive, workload_classes,
                       meta=None, extra=None):
    """Assemble the machine-readable summary emitted to stdout under --json."""
    out = dict(meta or {})
    out["scopes"] = [
        {"id": s["scope_id"], "name": s["scope_name"], "workbook": s["workbook"],
         "workloads": _workload_totals(s["data"], workload_classes)}
        for s in scope_items
    ]
    out["combined"] = _workload_totals(combined, workload_classes)
    out["comprehensive_workbook"] = comprehensive
    if extra and extra.get("rows"):
        out[extra.get("key", "extra_summary")] = {lbl: val for lbl, val in extra["rows"]}
    return out


def print_banner(title, sessions, workloads, version="", subtitle="",
                 scope_label="Scopes"):
    """Concise startup summary of what is about to run (-> stderr)."""
    scopes = ", ".join(
        (f"{sname} ({sid or '?'})" if sname else f"{sid or '?'}")
        for _handle, sid, sname in sessions)
    head = f"[bold]{title}[/bold]" + (f" v{version}" if version else "")
    if subtitle:
        head += f"  -  {subtitle}"
    console.print(head)
    console.print()
    console.print(f"{scope_label}: {scopes}")
    console.print(f"Scanning for: {', '.join(workloads)}")
    console.print()


# --------------------------------------------------------------------------- #
# Shared credential-failure helper
# --------------------------------------------------------------------------- #
def dominant_failure_kind(failures, preferred=(
        "profile_not_found", "expired", "invalid_key",
        "access_denied", "no_credentials")):
    """Pick the most informative error kind across failed credential checks.

    `failures` is a list of (label, kind, detail). The default `preferred`
    order is a superset across clouds (AWS can emit profile_not_found; Azure
    never does, so it is simply skipped there)."""
    kinds = [k for _, k, _ in failures]
    for p in preferred:
        if p in kinds:
            return p
    return kinds[0] if kinds else "no_credentials"


# --------------------------------------------------------------------------- #
# Shared CLI argument helpers
# --------------------------------------------------------------------------- #
def _csv(value):
    """argparse type: split a comma-separated value into a clean list."""
    return [x.strip() for x in value.split(",") if x.strip()]


def add_common_args(parser, version, workload_classes):
    """Add the cloud-agnostic flags (--version, --workload, output group)."""
    parser.add_argument("--version", action="version",
                        version=f"%(prog)s {version}")
    out = parser.add_argument_group("output")
    out.add_argument("--workload", default="all", metavar="w1,w2",
                     help="workloads to collect (default: all). choices: "
                          + ", ".join(cls.WORKLOAD for cls in workload_classes)
                          + ", all")
    out.add_argument("--validate-only", action="store_true",
                     help="check credentials and print what would be collected, then exit")
    out.add_argument("--json", action="store_true",
                     help="emit a machine-readable summary to stdout")
    out.add_argument("-q", "--quiet", action="store_true",
                     help="only warnings and errors; no progress or summary")
    out.add_argument("-v", "--verbose", action="count", default=0,
                     help="verbose console output (timestamps, debug, tracebacks)")
    out.add_argument("--no-color", action="store_true",
                     help="disable colored output (also honors NO_COLOR)")
    out.add_argument("--no-input", action="store_true",
                     help="never prompt; for non-interactive / CI use")


def resolve_workloads(workload_arg, workload_classes):
    """Resolve a --workload value into a validated list of workload keys."""
    keys = [cls.WORKLOAD for cls in workload_classes]
    if not workload_arg or workload_arg == "all":
        return list(keys)
    requested = [w.strip() for w in workload_arg.split(",") if w.strip()]
    valid = set(keys)
    out = []
    for w in requested:
        if w not in valid:
            console.print(f"Unsupported workload '{w}'. Valid: "
                          f"{', '.join(keys)}, all", style="red")
            sys.exit(2)
        out.append(w)
    return out


# --------------------------------------------------------------------------- #
# Interactive scope picker (subscriptions / accounts)
# --------------------------------------------------------------------------- #
def _parse_selection(resp, n):
    """Parse '1,3-4 6' into a sorted, de-duped list of 1-based indices in 1..n.

    Returns None if any token is non-numeric, malformed, or out of range.
    """
    chosen = set()
    for token in resp.replace(",", " ").split():
        if "-" in token:
            lo, _, hi = token.partition("-")
            if not (lo.isdigit() and hi.isdigit()):
                return None
            lo, hi = int(lo), int(hi)
            if lo < 1 or hi > n or lo > hi:
                return None
            chosen.update(range(lo, hi + 1))
        else:
            if not token.isdigit():
                return None
            i = int(token)
            if i < 1 or i > n:
                return None
            chosen.add(i)
    return sorted(chosen) if chosen else None


def select_scopes(sessions, noun):
    """Prompt (on stderr) for which scopes to scan; return the chosen subset.

    The prompt renders to the shared stderr console and reads stdin via bare
    input(), so nothing leaks to stdout (which stays machine-only)."""
    # markup/highlight off so '[1]' and '[all]' render literally and the list
    # isn't auto-colorized into noise.
    console.print(f"\nDiscovered {len(sessions)} {noun}:", style="bold",
                  highlight=False)
    for i, (_handle, sid, sname) in enumerate(sessions, 1):
        console.print(f"  [{i}] {sname or sid}  ({sid})",
                      markup=False, highlight=False)
    console.print(f"  [a] all {noun}", markup=False, highlight=False)
    console.print(f"Select which {noun} to scan: numbers (e.g. 1,3-4), "
                  f"'a' for all, or 'q' to quit.", markup=False, highlight=False)
    while True:
        console.print("Selection [a]: ", end="", markup=False, highlight=False)
        try:
            resp = input().strip()
        except (EOFError, KeyboardInterrupt):
            console.print("\nCancelled.", style="yellow")
            sys.exit(130)
        if resp.lower() in ("q", "quit"):
            console.print("Cancelled.", style="yellow")
            sys.exit(0)
        if resp == "" or resp.lower() in ("a", "all"):
            return sessions
        idx = _parse_selection(resp, len(sessions))
        if idx is None:
            console.print(f"Invalid selection; enter numbers 1-{len(sessions)}, "
                          f"'a' for all, or 'q'.", style="red", markup=False,
                          highlight=False)
            continue
        chosen = [sessions[i - 1] for i in idx]
        names = ", ".join(sname or sid for _, sid, sname in chosen)
        console.print(f"Selected: {names}", style="green", markup=False,
                      highlight=False)
        return chosen


def _maybe_select_scopes(config, args, sessions):
    """Show the interactive picker only when appropriate; else pass through."""
    if (args.get("no_input") or args.get("quiet") or args.get("validate_only")
            or len(sessions) <= 1
            or not (sys.stdin.isatty() and console.is_terminal)):
        return sessions
    if config.scope_already_filtered and config.scope_already_filtered(args):
        return sessions
    return select_scopes(sessions, f"{config.scope_noun}s")


# --------------------------------------------------------------------------- #
# Driver: owns the main() control flow + entrypoint error handling.
# --------------------------------------------------------------------------- #
@dataclass
class CloudConfig:
    """Cloud-specific hooks + metadata consumed by run_sizing().

    args dict (returned by parse_args) must contain at least these keys consumed
    by the driver: workload, validate_only, json, quiet, verbose, no_color,
    no_input (plus whatever the cloud's own hooks need).
    """
    name: str                  # "AWS" / "Azure"
    version: str
    log_prefix: str            # "aws_sizing" / "azure_sizing"
    comprehensive_label: str   # "all_aws_accounts" / "all_azure_subscriptions"
    scope_noun: str            # "account" / "subscription"
    unit_noun: str             # "regions" / "workloads"
    workload_classes: list
    parse_args: object         # callable(argv) -> args dict (+ side effects)
    validate_args: object      # callable(args) -> None (may sys.exit)
    build_sessions: object     # callable(args) -> (sessions, failures)
    collect: object            # callable(handle, scope_id, scope_name, args,
    #                            workloads, reporter, label) -> data dict
    describe_scope: object     # callable(handle, args) -> str  (validate-only line)
    print_credential_help: object  # callable(failures) -> bool (reconfigured?)
    noisy_loggers: tuple = ()
    banner_subtitle: object = None  # optional callable(args) -> str
    scope_already_filtered: object = None  # callable(args) -> bool (skip picker)
    # optional callable(combined) -> {"title", "key", "rows": [(label, value)]} | None
    # for an extra section in the summary table + --json (e.g. backup coverage)
    extra_summary_rows: object = None


def run_sizing(config):
    """Top-level entrypoint: run the sizing flow with interrupt/error handling."""
    try:
        _run(config)
    except KeyboardInterrupt:
        console.print("\nInterrupted. Any workbooks already written remain in "
                      "Metrics/.", style="yellow")
        sys.exit(130)
    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001 - top-level safety net
        if _VERBOSE:
            console.print_exception()
        else:
            console.print(f"Unexpected error: {exc}", style="red")
            console.print("Re-run with --verbose for a full traceback.", style="dim")
        sys.exit(1)


def _run(config):
    global _VERBOSE
    args = config.parse_args(sys.argv[1:])
    _VERBOSE = bool(args.get("verbose"))
    quiet = bool(args.get("quiet"))

    if args.get("no_color"):
        disable_color()

    workloads = resolve_workloads(args.get("workload"), config.workload_classes)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    os.makedirs("Metrics", exist_ok=True)
    setup_logging(timestamp, log_prefix=config.log_prefix, quiet=quiet,
                  verbose=_VERBOSE, noisy_loggers=config.noisy_loggers)

    config.validate_args(args)

    # Credential preflight: only authenticated scopes proceed.
    sessions, failures = config.build_sessions(args)
    for label, _kind, detail in failures:
        if sessions:  # partial run -> warn and keep going with the valid ones
            logging.warning(f"Skipping '{label}': {detail}")

    if not sessions:
        # Zero authenticated scopes: print guidance and stop. If the user
        # reconfigures interactively, retry once with the same args.
        if config.print_credential_help(failures):
            sessions, _ = config.build_sessions(args)
        if not sessions:
            logging.error(f"No {config.name} credentials could be authenticated. "
                          f"Exiting.")
            sys.exit(1)

    # Interactive scope picker: when run on a TTY without an explicit scope
    # filter, let the user choose which subscriptions/accounts to scan.
    sessions = _maybe_select_scopes(config, args, sessions)

    if not quiet:
        subtitle = config.banner_subtitle(args) if config.banner_subtitle else ""
        print_banner(
            f"{config.name} Cloud Sizing",
            sessions,
            workloads,
            version=config.version,
            subtitle=subtitle,
            scope_label=f"{config.scope_noun.capitalize()}s",
        )

    if args.get("validate_only"):
        for handle, scope_id, scope_name in sessions:
            console.print(f"  {scope_id or '(unknown)'} ({scope_name or 'no name'})"
                          f" - {config.describe_scope(handle, args)}", style="green")
        console.print("Credentials valid; --validate-only set, not collecting.",
                      style="green")
        return

    combined = {cls.WORKLOAD: [] for cls in config.workload_classes}
    per_scope = []
    written = []

    with Reporter(enabled=not quiet, unit=config.unit_noun) as reporter:
        for handle, scope_id, scope_name in sessions:
            label = scope_name or scope_id or "unknown"
            logging.debug(f"=== Collecting {config.scope_noun} {scope_id} "
                          f"({scope_name or 'no name'}) ===")
            data = config.collect(handle, scope_id, scope_name, args, workloads,
                                  reporter, label)

            workbook = None
            if any(data.get(cls.WORKLOAD) for cls in config.workload_classes):
                safe = re.sub(r"[^A-Za-z0-9_.-]", "_", str(scope_id or label))
                workbook = os.path.join("Metrics", f"{safe}_summary_{timestamp}.xlsx")
                write_workbook(workbook, data, config.workload_classes)
                written.append(workbook)
                if not quiet:
                    console.print(f"Wrote {workbook}", style="green")
            else:
                logging.warning(f"No resources found for {label}; skipping workbook.")

            per_scope.append({"scope_id": scope_id, "scope_name": scope_name,
                              "workbook": workbook, "data": data})
            for cls in config.workload_classes:
                combined[cls.WORKLOAD].extend(data.get(cls.WORKLOAD, []))

    comprehensive = None
    if len(sessions) > 1 and any(combined[cls.WORKLOAD]
                                 for cls in config.workload_classes):
        comprehensive = os.path.join(
            "Metrics", f"comprehensive_{config.comprehensive_label}_{timestamp}.xlsx")
        write_workbook(comprehensive, combined, config.workload_classes)
        written.append(comprehensive)
        if not quiet:
            console.print(f"Wrote {comprehensive}", style="green")

    # Full grand totals always go to the file log for analysts.
    for cls in config.workload_classes:
        infos = combined[cls.WORKLOAD]
        if infos:
            _, grand = build_summary(infos, cls)
            logging.debug(f"Grand total - {cls.INFO_SHEET}: {grand}")

    extra_summary = (config.extra_summary_rows(combined)
                     if getattr(config, "extra_summary_rows", None) else None)

    if not quiet:
        render_summary_table(combined, config.workload_classes, extra_summary)

    # Machine-readable output -> stdout (and ONLY this goes to stdout).
    if args.get("json"):
        meta = {"version": config.version, "cloud": config.name.lower()}
        if config.banner_subtitle:
            meta["scope_note"] = config.banner_subtitle(args)
        print(json.dumps(build_json_summary(per_scope, combined, comprehensive,
                                            config.workload_classes, meta,
                                            extra_summary), indent=2))
    else:
        for path in written:
            print(path)

    logging.info("Done.")
